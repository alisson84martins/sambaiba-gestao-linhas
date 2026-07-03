"""
Upload de escala gerencial (.xlsx) → popula partidas_programadas.

Formato esperado: abas com nome no padrão XXXX-YY (ex: 1726-10).
Cada aba tem: cabeçalho TABELAS na linha de títulos, colunas = números de tabela,
linhas = horários e duplas (par mot/cob).
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from datetime import date, time
from typing import Optional
import openpyxl
import re
import io

from app.core.database import get_db
from app.core.deps import get_current_usuario
from app.models.fiscalizacao import PartidaProgramada, LinhaFiscal, PrevisaoRecolhida
from app.models.v3 import UsuarioV3

router = APIRouter(prefix="/escalas", tags=["Escalas"])

# Regex do nome de aba: 4-5 alfanum + hífen + 2 dígitos
_RE_ABA = re.compile(r'^[0-9A-Z]{4,5}-\d{2}$')

def _serial_to_time(val) -> Optional[time]:
    """Converte valor Excel em time. Aceita datetime.time ou fração de dia (0–1)."""
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, float):
        if val >= 1.0:
            return None  # data/número (km, horas trabalhadas), não horário
        total_minutes = round(val * 24 * 60)
        h, m = divmod(total_minutes, 60)
        return time(h % 24, m)
    return None


def _parse_range_second(val: str) -> Optional[time]:
    """
    Extrai o SEGUNDO horário de uma string 'HH:MM-HH:MM'.
    Usado para células de refeição: '07:55-08:55' → 08:55 (partida após refeição).
    """
    try:
        partes = val.strip().split("-")
        if len(partes) == 2:
            h_m = partes[1].strip().split(":")
            if len(h_m) == 2:
                return time(int(h_m[0]) % 24, int(h_m[1]))
    except (ValueError, IndexError):
        pass
    return None


def _is_ts_label(col_a) -> bool:
    """True se o label da coluna A indica terminal TS (Metrô / terminal secundário)."""
    if not isinstance(col_a, str):
        return False
    label = col_a.strip().upper()
    return "METR" in label or label == "TS"


def _parse_aba(ws) -> list[dict]:
    """
    Parseia uma aba da escala Sambaíba.

    Estrutura real das linhas de dados:
      HORÁRIOS      | 04:15 | ...  ← saída da garagem, IGNORAR
      None          | 04:15 | ...  ← linha em branco / duplicata, IGNORAR
      CEM PQ (TP)   | 04:30 | ...  ← partida do TP   ┐ par
      METRÔ (TS)    | 05:20 | ...  ← partida do TS   ┘
      CEM PQ (TP)   | '07:55-08:55' | ...  ← chegada 07:55, saída 08:55 (refeição) → usar 2º horário
      METRÔ (TS)    | 10:00 | ...
      CEM PQ (TP)   | 13:10 | ...  ← CHEGADA ao TP (METRÔ seguinte tem 'TERM.') → IGNORAR
      METRÔ (TS)    | 'TERM.' | ... ← marcador de fim de período
      DUPLAS        | ...          ← seção de informações — PARAR aqui

    Regra para o par (CEM PQ, METRÔ):
      - Se METRÔ da coluna tem horário real → CEM PQ é PARTIDA TP + METRÔ é PARTIDA TS
      - Se METRÔ tem string ('TERM.', 'REC.', 'INICIO') ou None → CEM PQ era CHEGADA → ignorar par
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 1. Localiza linha onde col A == "TABELAS"
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().upper() == "TABELAS":
            header_row_idx = i
            break

    if header_row_idx is None:
        return []

    # 2. Mapeia colunas → números de tabela (inteiros 1–30)
    tabela_cols: dict[int, int] = {}
    for j, cell in enumerate(rows[header_row_idx]):
        if j == 0:
            continue
        if isinstance(cell, (int, float)) and 1 <= int(cell) <= 30:
            tabela_cols[j] = int(cell)

    if not tabela_cols:
        return []

    # 3. Coleta linhas de dados relevantes (pula HORÁRIOS e None; para em DUPLAS)
    data_rows: list = []
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        col_a = row[0]
        # Seção DUPLAS = informações de escala, não horários
        if isinstance(col_a, str) and "DUPLAS" in col_a.strip().upper():
            break
        # Ignora linhas sem label (entre seções) e HORÁRIOS (saída da garagem)
        if col_a is None:
            continue
        if isinstance(col_a, str) and col_a.strip().upper() == "HORÁRIOS":
            continue
        data_rows.append(row)

    # 4. Processa pares (TP, TS) em sequência
    partidas: list[dict] = []
    seq_tp: dict[int, int] = {t: 0 for t in tabela_cols.values()}
    seq_ts: dict[int, int] = {t: 0 for t in tabela_cols.values()}

    i = 0
    while i < len(data_rows):
        row = data_rows[i]
        col_a = row[0]

        if not isinstance(col_a, str):
            i += 1
            continue

        if not _is_ts_label(col_a):
            # Linha TP — consome este e o próximo (TS)
            ts_row = data_rows[i + 1] if i + 1 < len(data_rows) else None

            for col_idx, num_tabela in tabela_cols.items():
                tp_cell = row[col_idx] if col_idx < len(row) else None
                ts_cell = ts_row[col_idx] if (ts_row and col_idx < len(ts_row)) else None

                # Verifica se o TS deste par tem horário real
                ts_time = _serial_to_time(ts_cell)
                if ts_time is None:
                    # TS tem string marcador ('TERM.', 'REC.', 'INICIO') ou None
                    # → TP era chegada, não partida → ignora o par
                    continue

                # TS tem horário → TP é partida real
                # Determina horário TP (pode ser range "HH:MM-HH:MM" para refeição)
                tp_time = _serial_to_time(tp_cell)
                if tp_time is None and isinstance(tp_cell, str) and "-" in tp_cell:
                    tp_time = _parse_range_second(tp_cell)  # usa horário pós-refeição

                if tp_time is not None:
                    seq_tp[num_tabela] += 1
                    partidas.append({
                        "numero_tabela": num_tabela,
                        "sequencia": seq_tp[num_tabela],
                        "horario": tp_time,
                        "terminal": "TP",
                        "mot_re": None,
                        "cob_re": None,
                    })

                # Partida TS
                seq_ts[num_tabela] += 1
                partidas.append({
                    "numero_tabela": num_tabela,
                    "sequencia": seq_ts[num_tabela],
                    "horario": ts_time,
                    "terminal": "TS",
                    "mot_re": None,
                    "cob_re": None,
                })

            i += 2  # consome TP + TS

        else:
            # Linha TS sem TP precedente (não deve ocorrer normalmente)
            i += 1

    return partidas


def _parse_previsao_recolhida(ws) -> list[dict]:
    """
    Best-effort: captura o horário previsto de chegada na garagem (marcador 'REC.').

    Quando a célula do terminal TS (METRÔ) de uma tabela é 'REC.', a célula da
    MESMA coluna na linha seguinte do grid contém o horário previsto de recolhida
    (mesmo padrão observado na linha resumo 'RECOLHE' da seção de duplas).
    Roda isolado de _parse_aba (não compartilha estado) — se a estrutura da aba
    não seguir exatamente esse padrão, o item correspondente é apenas ignorado,
    sem travar o upload. Dados não capturados ficam para preenchimento manual
    via POST /coordenador/previsao-recolhida.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_idx = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().upper() == "TABELAS":
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    tabela_cols: dict[int, int] = {}
    for j, cell in enumerate(rows[header_row_idx]):
        if j == 0:
            continue
        if isinstance(cell, (int, float)) and 1 <= int(cell) <= 30:
            tabela_cols[j] = int(cell)
    if not tabela_cols:
        return []

    data_rows: list = []
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        col_a = row[0]
        if isinstance(col_a, str) and "DUPLAS" in col_a.strip().upper():
            break
        if col_a is None:
            continue
        if isinstance(col_a, str) and col_a.strip().upper() == "HORÁRIOS":
            continue
        data_rows.append(row)

    # Rastreia o período (MANHA/TARDE) corrente de cada tabela — vira TARDE
    # após o marcador 'INICIO' (início da jornada da tarde), igual ao padrão
    # já usado pelo periodo_enum em turno_fiscal.
    periodo_atual = {t: "MANHA" for t in tabela_cols.values()}
    encontrados: dict[tuple, time] = {}

    for i, row in enumerate(data_rows):
        col_a = row[0]
        if not isinstance(col_a, str):
            continue

        for col_idx, num_tabela in tabela_cols.items():
            cell = row[col_idx] if col_idx < len(row) else None
            if isinstance(cell, str) and cell.strip().upper() == "INICIO":
                periodo_atual[num_tabela] = "TARDE"

        if not _is_ts_label(col_a):
            continue

        for col_idx, num_tabela in tabela_cols.items():
            cell = row[col_idx] if col_idx < len(row) else None
            if not (isinstance(cell, str) and cell.strip().upper() == "REC."):
                continue
            if i + 1 >= len(data_rows):
                continue
            prox_row = data_rows[i + 1]
            valor = prox_row[col_idx] if col_idx < len(prox_row) else None
            horario = _serial_to_time(valor)
            if horario is None:
                continue
            encontrados[(num_tabela, periodo_atual[num_tabela])] = horario

    return [
        {"numero_tabela": t, "periodo": p, "horario_previsto": h}
        for (t, p), h in encontrados.items()
    ]


@router.post("/upload")
async def upload_escala(
    file: UploadFile = File(...),
    tipo_dia: str = "UTIL",
    vigencia: date = None,
    usuario: UsuarioV3 = Depends(get_current_usuario),
    db: AsyncSession = Depends(get_db),
):
    """
    Recebe um .xlsx de escala gerencial.
    Para cada aba com nome no padrão XXXX-YY:
      1. Cria/confirma LinhaFiscal
      2. Remove partidas antigas do mesmo tipo_dia + vigência
      3. Insere novas PartidaProgramada
    """
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx são aceitos")

    if vigencia is None:
        vigencia = date.today()

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    result_summary = []

    for sheet_name in wb.sheetnames:
        codigo = sheet_name.strip().upper().replace(" ", "")
        if not _RE_ABA.match(codigo):
            continue

        ws = wb[sheet_name]
        partidas = _parse_aba(ws)
        if not partidas:
            continue

        # Garante LinhaFiscal existe
        linha_result = await db.execute(
            select(LinhaFiscal).where(LinhaFiscal.codigo == codigo)
        )
        linha = linha_result.scalar_one_or_none()
        if not linha:
            linha = LinhaFiscal(codigo=codigo, nome=f"Linha {codigo}")
            db.add(linha)
            await db.flush()

        # Remove partidas antigas deste tipo_dia + vigência para esta linha
        await db.execute(
            delete(PartidaProgramada).where(
                PartidaProgramada.linha_codigo == codigo,
                PartidaProgramada.tipo_dia == tipo_dia,
                PartidaProgramada.vigencia == vigencia,
            )
        )
        await db.flush()  # garante que o DELETE é enviado ao BD antes do INSERT

        # Insere novas
        for p in partidas:
            pp = PartidaProgramada(
                linha_id=linha.id,
                linha_codigo=codigo,
                tipo_dia=tipo_dia,
                numero_tabela=p["numero_tabela"],
                sequencia=p["sequencia"],
                terminal=p["terminal"],
                horario=p["horario"],
                vigencia=vigencia,
            )
            db.add(pp)

        # Horário previsto de recolhida (best-effort — ver _parse_previsao_recolhida)
        previsoes_rec = _parse_previsao_recolhida(ws)
        if previsoes_rec:
            await db.execute(
                delete(PrevisaoRecolhida).where(
                    PrevisaoRecolhida.linha_codigo == codigo,
                    PrevisaoRecolhida.tipo_dia == tipo_dia,
                    PrevisaoRecolhida.vigencia == vigencia,
                )
            )
            await db.flush()
            for pr in previsoes_rec:
                db.add(PrevisaoRecolhida(
                    linha_codigo=codigo,
                    tipo_dia=tipo_dia,
                    numero_tabela=pr["numero_tabela"],
                    periodo=pr["periodo"],
                    horario_previsto=pr["horario_previsto"],
                    vigencia=vigencia,
                ))

        result_summary.append({
            "linha": codigo,
            "partidas_importadas": len(partidas),
            "previsoes_recolhida_capturadas": len(previsoes_rec),
        })

    await db.commit()
    return {
        "ok": True,
        "vigencia": str(vigencia),
        "tipo_dia": tipo_dia,
        "linhas": result_summary,
    }


@router.get("/partidas/{linha_codigo}")
async def listar_partidas(
    linha_codigo: str,
    tipo_dia: str = "UTIL",
    terminal: Optional[str] = None,
    usuario: UsuarioV3 = Depends(get_current_usuario),
    db: AsyncSession = Depends(get_db),
):
    """Retorna partidas programadas agrupadas por tabela, filtradas por terminal."""
    from sqlalchemy import and_

    conditions = [
        PartidaProgramada.linha_codigo == linha_codigo.upper(),
        PartidaProgramada.tipo_dia == tipo_dia,
    ]
    if terminal:
        conditions.append(PartidaProgramada.terminal == terminal.upper())

    result = await db.execute(
        select(PartidaProgramada).where(and_(*conditions))
        .order_by(PartidaProgramada.horario, PartidaProgramada.terminal)
    )
    partidas = result.scalars().all()

    # Agrupa por tabela
    tabelas: dict[int, list] = {}
    for p in partidas:
        tabelas.setdefault(p.numero_tabela, []).append({
            "id": str(p.id),
            "sequencia": p.sequencia,
            "horario": p.horario.strftime("%H:%M"),
            "terminal": p.terminal,
        })

    return [{"tabela": k, "partidas": v} for k, v in sorted(tabelas.items())]
